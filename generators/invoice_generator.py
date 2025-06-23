# invoice_generator.py

import io
from openpyxl import load_workbook
from datetime import date


def generate_invoice_file(fields: dict):
    wb = load_workbook("templates/invoice.xlsx")
    ws = wb.active

    ws["E3"] = fields["invoice_date"]
    ws["E4"] = f"{fields['case_id']}-{fields['invoice_number']}"
    ws["E5"] = fields["case_id"]
    ws["A10"] = fields["client_name"]
    ws["A11"] = fields["address"]
    ws["D17"] = fields["fd_letter_date"]

    ar_fee = float(fields.get("ar_fee") or 2)

    if fields["part_type"] == "Part B":
        ws["A17"] = "U.S. Department of Labor Part B Award"
        ws["B17"] = 150000
        ws["E16"] = f"{ar_fee}% AR Fee"
    elif fields["part_type"] == "Part E":
        ws["A17"] = "U.S. Department of Labor Part E"
        ws["B17"] = float(fields["awarded_amount"])
    ws["E17"] = ws["B17"].value * (ar_fee / 100)

    # invoicing
    amount_map = {
        "toupin": 200,
        "herold": 200,
        "klepper": 300,
        "smith": 300,
    }
    start_row = 20
    for i, item in enumerate(fields.get("invoice_items", [])[:8]):
        service_name = item["name"]
        service_date = item["date"]
        lower_name = service_name.lower()
        amount = next(
            (amount for key, amount in amount_map.items() if key in lower_name), ""
        )
        row = start_row + i
        ws[f"A{row}"] = service_name
        ws[f"B{row}"] = amount
        ws[f"E{row}"] = amount
        ws[f"D{row}"] = service_date

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Format amount
    amount = (
        float(fields["awarded_amount"]) if fields["part_type"] == "Part E" else 150000
    )
    amount_str = (
        f"${amount/1000:.1f}k" if amount % 1000 != 0 else f"${int(amount/1000)}k"
    )

    # Extract client short name (assumes format: "First Last")
    short_name = fields["client_name"].split()[-1]  # Last name only
    first_initial = fields["client_name"][0]  # First initial
    name_str = f"{first_initial}. {short_name}"

    # Get today's date
    today_str = date.today().strftime("%m.%d.%y")

    filename = f"Invoice {fields['part_type']} {amount_str} {name_str} {today_str}.xlsx"
    return filename, buffer
